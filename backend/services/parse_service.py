"""Service layer wrapping the skill's parse_data.py functions."""
import sys
import os
import json
import re
import uuid

# Read-only import from the skill scripts (bundled or local)
from config import SKILL_SCRIPTS_DIR, UPLOAD_DIR
if SKILL_SCRIPTS_DIR not in sys.path:
    sys.path.insert(0, SKILL_SCRIPTS_DIR)

import parse_data


def parse_file(file_id: str, column_map: dict = None, daily: bool = False) -> dict:
    """Parse an uploaded file by file_id."""
    # Find the file in uploads
    filepath = None
    for fname in os.listdir(UPLOAD_DIR):
        if fname.startswith(file_id):
            filepath = os.path.join(UPLOAD_DIR, fname)
            break

    if not filepath:
        return {"error": f"文件不存在或已过期: {file_id}", "building_info": {}, "energy_data": []}

    ext = os.path.splitext(filepath)[1].lower()

    if column_map:
        result = parse_data.parse_with_columns(filepath, column_map, daily=daily)
    elif ext in ('.xlsx', '.xls'):
        result = parse_data.parse_excel(filepath)
    elif ext == '.csv':
        result = parse_data.parse_csv_file(filepath)
    else:
        result = parse_data.parse_table_text(open(filepath, encoding='utf-8').read())

    # Add data quality validations
    energy_data = result.get('energy_data', [])
    if energy_data:
        try:
            import energy_analysis
            warnings = energy_analysis.validate_data(energy_data)
            result['warnings'] = warnings
        except Exception:
            result['warnings'] = []

    return result


def parse_text(raw_text: str) -> dict:
    """Parse pasted table text."""
    result = parse_data.parse_table_text(raw_text)

    energy_data = result.get('energy_data', [])
    if energy_data:
        try:
            import energy_analysis
            warnings = energy_analysis.validate_data(energy_data)
            result['warnings'] = warnings
        except Exception:
            result['warnings'] = []

    return result


def preview_file(filepath: str, max_rows: int = 5):
    """Preview an Excel/CSV file structure."""
    ext = os.path.splitext(filepath)[1].lower()
    if ext in ('.xlsx', '.xls'):
        return parse_data.preview_excel(filepath, max_rows=max_rows)
    elif ext == '.csv':
        rows = []
        import csv
        with open(filepath, 'r', encoding='utf-8-sig') as f:
            reader = csv.reader(f)
            for i, row in enumerate(reader):
                if i >= max_rows:
                    break
                rows.append(row)
        return {'rows': rows, 'total_columns': len(rows[0]) if rows else 0}
    return {}


def parse_file_transposed(file_id: str, sheet_name: str = None,
                          start_row: int = 2, month_start_col: int = 2,
                          num_months: int = 12, year: int = 2025) -> dict:
    """
    Parse transposed data: rows = buildings/sub-items, columns = months.
    Sums across all rows to get monthly totals.

    Args:
        file_id: Uploaded file ID
        sheet_name: Sheet to parse (None = auto-detect)
        start_row: First data row (1-based, after header)
        month_start_col: First month column (1-based)
        num_months: Number of month columns
        year: Data year
    """
    import openpyxl

    filepath = None
    for fname in os.listdir(UPLOAD_DIR):
        if fname.startswith(file_id):
            filepath = os.path.join(UPLOAD_DIR, fname)
            break

    if not filepath:
        return {"error": f"文件不存在或已过期: {file_id}", "building_info": {}, "energy_data": []}

    wb = openpyxl.load_workbook(filepath, data_only=True)

    # Select sheet
    ws = None
    if sheet_name:
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
    if ws is None:
        # Auto-detect: prefer sheets with energy keywords
        energy_keywords = ['用能', '用电', '能耗', '能源', '电', '气']
        for sn in wb.sheetnames:
            if any(kw in sn for kw in energy_keywords):
                ws = wb[sn]
                break
        if ws is None:
            ws = wb.active

    sheet_name = ws.title

    # Read all data rows and sum by column
    monthly_totals = {m: 0.0 for m in range(1, num_months + 1)}
    row_count = 0

    for ri in range(start_row, ws.max_row + 1):
        first_cell = ws.cell(row=ri, column=1).value
        if not first_cell:
            continue  # Skip empty rows

        for mi in range(num_months):
            col = month_start_col + mi
            val = ws.cell(row=ri, column=col).value
            if val is not None:
                try:
                    monthly_totals[mi + 1] += float(str(val).replace(',', ''))
                except (ValueError, TypeError):
                    pass
        row_count += 1

    wb.close()

    # Build energy_data in standard format
    energy_data = []
    for month in range(1, num_months + 1):
        energy_data.append({
            'month': month,
            'electricity_kwh': round(monthly_totals[month], 2),
            'gas_m3': 0,
            'heat_gj': 0,
        })

    # Try to extract building info
    building_info = {"name": "", "area": 0, "type": "", "location": "", "year": year, "source": filepath}

    # Try to detect name from filename
    basename = os.path.splitext(os.path.basename(filepath))[0]
    if basename and not building_info['name']:
        building_info['name'] = basename

    return {
        "building_info": building_info,
        "sheet_name": sheet_name,
        "data_type": "transposed_monthly",
        "energy_data": energy_data,
        "warnings": [],
        "meta": {"rows_aggregated": row_count, "months": num_months},
    }


def get_file_preview(file_id: str, max_rows: int = 10) -> dict:
    """Get a detailed preview of an uploaded file's structure for UI display."""
    filepath = None
    for fname in os.listdir(UPLOAD_DIR):
        if fname.startswith(file_id):
            filepath = os.path.join(UPLOAD_DIR, fname)
            break

    if not filepath:
        return {"error": "文件不存在或已过期"}

    ext = os.path.splitext(filepath)[1].lower()
    if ext not in ('.xlsx', '.xls'):
        return {"error": "预览仅支持 Excel 文件"}

    import openpyxl
    wb = openpyxl.load_workbook(filepath, data_only=True)

    sheets = []
    for sn in wb.sheetnames:
        ws = wb[sn]
        rows = []
        col_names = []  # Column index -> first cell value (for mapping)

        for ri in range(1, min(max_rows + 1, ws.max_row + 1)):
            cells = {}
            for ci in range(1, min(ws.max_column + 1, 50)):
                val = ws.cell(row=ri, column=ci).value
                if val is not None:
                    # Excel serial date conversion
                    if isinstance(val, (int, float)) and 40000 < val < 60000 and ci > 1:
                        import datetime
                        try:
                            dt = datetime.datetime(1899, 12, 30) + datetime.timedelta(days=int(val))
                            val_str = dt.strftime('%Y-%m-%d')
                        except Exception:
                            val_str = str(val)
                    else:
                        val_str = str(val)[:50]
                    cells[str(ci)] = val_str
            if cells:
                rows.append({"row": ri, "cells": cells})

        # Get column headers from first row for mapping suggestions
        first_row = rows[0]["cells"] if rows else {}
        col_headers = {k: v for k, v in first_row.items()}

        sheets.append({
            "name": sn,
            "total_rows": ws.max_row,
            "total_cols": ws.max_column,
            "preview_rows": rows,
            "headers": col_headers,
        })

    wb.close()

    # Detect likely data structure
    detection = _detect_structure(sheets)

    return {
        "sheets": sheets,
        "detection": detection,
    }


def _detect_structure(sheets: list) -> dict:
    """Try to detect whether data is standard (rows=months) or transposed (rows=buildings)."""
    result = {"mode": "standard", "confidence": 0, "suggestion": ""}

    for sheet in sheets:
        rows = sheet.get("preview_rows", [])
        if len(rows) < 2:
            continue

        # Check first column of data rows
        first_col_values = []
        for r in rows[1:]:
            cells = r.get("cells", {})
            if "1" in cells:
                first_col_values.append(cells["1"])

        # If first column has text values (building names), it's likely transposed
        text_count = sum(1 for v in first_col_values if v and not v.replace('.', '').replace('-', '').isdigit())
        if text_count > len(first_col_values) * 0.5 and len(first_col_values) >= 3:
            result = {
                "mode": "transposed",
                "confidence": min(90, text_count * 25),
                "suggestion": "检测到每行数据对应一个建筑/子项，每列为一个月。建议使用「按建筑汇总」模式解析（将自动汇总所有行的月度数据）。",
                "sheet": sheet["name"],
                "data_start_row": 2,
                "month_start_col": 2,
                "num_month_cols": sheet["total_cols"] - 1,
            }
            break

    return result


def save_upload_file(filename: str, content: bytes) -> tuple:
    """Save an uploaded file and return (file_id, filepath)."""
    file_id = str(uuid.uuid4())
    ext = os.path.splitext(filename)[1].lower() or '.xlsx'
    filepath = os.path.join(UPLOAD_DIR, f"{file_id}{ext}")
    with open(filepath, 'wb') as f:
        f.write(content)
    return file_id, filepath
