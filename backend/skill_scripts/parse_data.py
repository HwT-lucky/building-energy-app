"""
建筑能耗数据解析脚本 v2
支持:
  - Excel (.xlsx), CSV, Markdown 表格, 纯文本表格
  - 月度汇总表 + 每日累计抄表数据
  - 自动识别列类型（电/气/水/热）、乘数、初始读数
输出: 标准 JSON 到 stdout，每月一行
"""
import sys, json, os, re, csv
from collections import defaultdict
from datetime import datetime


# ============================================================
# Column type detection patterns (Chinese labels)
# ============================================================
COLUMN_PATTERNS = {
    'electricity_kwh': [
        r'电[用量耗]', r'用电', r'电\d', r'电力', r'electricity',
    ],
    'gas_m3': [
        r'[燃天]气', r'用气', r'气[用量耗]', r'锅炉气', r'厨房气', r'gas',
    ],
    'heat_gj': [
        r'热[力用量耗]', r'供热', r'用热', r'heat', r'暖气',
    ],
    'water_ton': [
        r'[用生]水', r'水[用量耗]', r'自来水', r'water',
    ],
}

# Multiplier patterns
MULTIPLIER_PATTERNS = [
    (r'\*(\d+)K倍', 1000),
    (r'\*(\d+)k倍', 1000),
    (r'\*(\d+)倍', 1),
    (r'[×xX*](\d+)K', 1000),
    (r'[×xX*](\d+)k', 1000),
    (r'[×xX*](\d+)', 1),
]

# Date column patterns
DATE_PATTERNS = [r'日期', r'时间', r'date', r'time', r'抄表日期']


def detect_input_type(input_str):
    """自动识别输入类型"""
    input_str = input_str.strip()
    if os.path.isfile(input_str):
        ext = os.path.splitext(input_str)[1].lower()
        if ext == '.xlsx':
            return 'excel', input_str
        elif ext == '.csv':
            return 'csv', input_str
        else:
            try:
                with open(input_str, 'r', encoding='utf-8') as f:
                    content = f.read()
                return 'text', content
            except Exception:
                return 'text', input_str
    return 'text', input_str


def detect_column_type(header_text):
    """Detect what type of energy data a column contains. Returns (type, multiplier)."""
    h = str(header_text).strip().lower()
    multiplier = 1

    # Check for multipliers
    for pattern, base in MULTIPLIER_PATTERNS:
        m = re.search(pattern, str(header_text))
        if m:
            multiplier = int(m.group(1)) * base
            break

    # Check column type
    for col_type, patterns in COLUMN_PATTERNS.items():
        for pat in patterns:
            if re.search(pat, h):
                return col_type, multiplier

    return None, 1


def is_date_column(header_text):
    """Check if column is a date column"""
    h = str(header_text).strip().lower()
    for pat in DATE_PATTERNS:
        if re.search(pat, h):
            return True
    return False


def is_header_row(row_values):
    """Determine if a row looks like a header row (text, not numbers)."""
    text_count = 0
    num_count = 0
    for v in row_values:
        if v is None:
            continue
        try:
            float(str(v).replace(',', ''))
            num_count += 1
        except (ValueError, TypeError):
            if str(v).strip():
                text_count += 1
    # Header rows have more text than numbers
    return text_count > num_count * 0.5 and text_count >= 2


def is_initial_readings_row(row_values, header_row_values):
    """Check if row contains initial meter readings (like Row 1 in daily data).
    Usually has all numeric values and the first cell is a date."""
    if not row_values:
        return False
    # First cell is a date
    first = row_values[0]
    if isinstance(first, datetime):
        pass  # OK this is likely data
    elif isinstance(first, str):
        # Check if it looks like a date
        try:
            datetime.strptime(str(first).strip()[:10], '%Y-%m-%d')
        except ValueError:
            return False
    else:
        try:
            float(str(first))
        except (ValueError, TypeError):
            return False

    # Most other cells are numeric
    num_count = sum(1 for v in row_values[1:] if v is not None and isinstance(v, (int, float)))
    return num_count >= len(row_values) * 0.3


def parse_excel(filepath):
    """Parse Excel file with advanced structure detection."""
    try:
        import openpyxl
    except ImportError:
        return {"error": "需要安装 openpyxl: pip install openpyxl", "building_info": {}, "energy_data": []}

    wb = openpyxl.load_workbook(filepath, data_only=True)

    # Check all sheets for energy data; prefer sheets with keywords
    energy_keywords = ['抄表', '能耗', '能源', '电', '气', 'energy', '用能', '走势', '统计']
    best_sheet = None
    for sname in wb.sheetnames:
        ws = wb[sname]
        # Quick check: does this sheet have enough data?
        if ws.max_row < 5:
            continue
        # Check first few cells for energy keywords
        first_text = ''
        for r in range(1, min(4, ws.max_row + 1)):
            for c in range(1, min(20, ws.max_column + 1)):
                v = ws.cell(row=r, column=c).value
                if isinstance(v, str):
                    first_text += v.lower()
        score = sum(1 for kw in energy_keywords if kw in first_text)
        if score > 0:
            best_sheet = sname
            break
        if best_sheet is None:
            best_sheet = sname  # Fallback to first sheet

    if best_sheet is None:
        best_sheet = wb.active.title

    return parse_excel_sheet(wb[best_sheet], filepath, best_sheet)


def parse_excel_sheet(ws, filepath, sheet_name):
    """Parse a single Excel sheet. Handles both monthly summary and daily cumulative data."""
    all_rows = []
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 1000),
                             max_col=min(ws.max_column, 100), values_only=True):
        cells = [c for c in row]
        if any(c is not None for c in cells):
            all_rows.append(cells)

    if len(all_rows) < 3:
        return {"error": f"Sheet '{sheet_name}' 数据不足", "building_info": {}, "energy_data": []}

    # ================================================================
    # Phase 1: Find header row
    # ================================================================
    header_row_idx = find_header_row_idx(all_rows)
    if header_row_idx is None:
        return {"error": f"Sheet '{sheet_name}': 无法识别表头行", "building_info": {}, "energy_data": []}

    headers = all_rows[header_row_idx]
    header_labels = [str(h) if h is not None else '' for h in headers]

    # Detect column types
    col_types = {}  # col_index -> (type, multiplier)
    col_metadata = {}  # ci -> {'is_consumption': bool, 'label': str}
    date_col = None
    for ci, h in enumerate(headers):
        if h is None:
            continue
        h_str = str(h).strip()
        if is_date_column(h):
            date_col = ci
        else:
            ctype, mult = detect_column_type(h)
            if ctype:
                col_types[ci] = (ctype, mult)
                col_metadata[ci] = {
                    'is_consumption': '用量' in h_str,
                    'label': h_str,
                }

    # Prefer "用量" (consumption) columns over cumulative meter columns
    # When both exist for the same type, keep only the consumption columns
    if col_metadata:
        for ctype in set(ct for ct, _ in col_types.values()):
            cols_of_type = [(ci, col_metadata.get(ci, {}).get('is_consumption', False))
                           for ci, (t, _) in col_types.items() if t == ctype]
            has_consumption = any(is_cons for _, is_cons in cols_of_type)
            if has_consumption:
                # Remove non-consumption columns of this type
                for ci, is_cons in cols_of_type:
                    if not is_cons:
                        del col_types[ci]

    if not col_types:
        return {"error": f"Sheet '{sheet_name}': 未识别到能耗数据列（电/气/热力）", "building_info": {}, "energy_data": []}

    # ================================================================
    # Phase 2: Extract building info from rows before header
    # ================================================================
    building_info = extract_building_info(all_rows[:header_row_idx], filepath)

    # ================================================================
    # Phase 3: Determine if daily cumulative or monthly summary
    # ================================================================
    data_start = header_row_idx + 1
    data_rows = all_rows[data_start:]

    # Check if Row 1 (before header) has initial cumulative readings
    initial_readings = {}
    if header_row_idx == 1:  # Row 0 = initial values, Row 1 = header, Row 2+ = data
        initial_row = all_rows[0]
        for ci in col_types:
            if ci < len(initial_row) and initial_row[ci] is not None:
                try:
                    initial_readings[ci] = float(str(initial_row[ci]).replace(',', ''))
                except (ValueError, TypeError):
                    pass

    # Detect if data is daily (many rows, dates are consecutive)
    is_daily = len(data_rows) > 24  # More than 24 rows = likely daily

    # ================================================================
    # Phase 4: Parse data rows
    # ================================================================
    if is_daily and initial_readings:
        # Daily cumulative meter data → calculate differences → aggregate monthly
        energy_data = process_daily_cumulative(
            data_rows, date_col, col_types, initial_readings, building_info)
    elif is_daily:
        # Daily but already has consumption values
        energy_data = process_daily_consumption(
            data_rows, date_col, col_types)
    else:
        # Monthly summary
        energy_data = process_monthly_data(data_rows, date_col, col_types)

    if not energy_data:
        return {"error": f"Sheet '{sheet_name}': 解析后无有效数据", "building_info": building_info, "energy_data": []}

    return {
        "building_info": building_info,
        "sheet_name": sheet_name,
        "data_type": "daily_cumulative" if (is_daily and initial_readings) else ("daily" if is_daily else "monthly"),
        "energy_data": energy_data,
    }


def find_header_row_idx(all_rows):
    """Find the row containing column headers."""
    for i, row in enumerate(all_rows):
        row_text = ' '.join(str(c) for c in row if c is not None).lower()
        if any(re.search(pat, row_text) for pat in [
            r'电[用量耗\d]', r'气[用量耗]', r'天然气', r'用[电气水]',
        ]):
            if is_header_row(row):
                return i
    # Fallback: first row with enough text
    for i, row in enumerate(all_rows):
        if is_header_row(row):
            return i
    return 0


def process_daily_cumulative(data_rows, date_col, col_types, initial_readings, building_info):
    """Process daily cumulative meter readings → monthly energy consumption."""
    monthly = defaultdict(lambda: defaultdict(float))

    prev_readings = dict(initial_readings)

    for row in data_rows:
        # Get date
        month = None
        if date_col is not None and date_col < len(row):
            dt_val = row[date_col]
            if isinstance(dt_val, datetime):
                if dt_val.year < 2000:  # Excel serial date weirdness
                    continue
                month = dt_val.month
            elif isinstance(dt_val, str):
                try:
                    dt = datetime.strptime(str(dt_val).strip()[:10], '%Y-%m-%d')
                    month = dt.month
                except ValueError:
                    try:
                        dt = datetime.strptime(str(dt_val).strip()[:10], '%Y/%m/%d')
                        month = dt.month
                    except ValueError:
                        continue
            elif isinstance(dt_val, (int, float)):
                # Excel serial number - skip for now
                continue

        if month is None:
            continue

        # Calculate consumption = current - previous
        for ci, (ctype, mult) in col_types.items():
            if ci >= len(row) or row[ci] is None:
                continue
            try:
                current = float(str(row[ci]).replace(',', ''))
            except (ValueError, TypeError):
                continue

            prev = prev_readings.get(ci, current)
            consumption = current - prev
            if consumption < 0:
                # Meter rollover - use 0 for this day
                consumption = 0

            # Apply multiplier
            monthly[month][ctype] += consumption * mult

            # Update previous reading
            prev_readings[ci] = current

    return build_energy_data(monthly)


def process_daily_consumption(data_rows, date_col, col_types):
    """Process daily consumption values (not cumulative) → monthly."""
    monthly = defaultdict(lambda: defaultdict(float))

    for row in data_rows:
        month = None
        if date_col is not None and date_col < len(row):
            dt_val = row[date_col]
            if isinstance(dt_val, datetime) and dt_val.year >= 2000:
                month = dt_val.month
            elif isinstance(dt_val, str):
                try:
                    month = datetime.strptime(str(dt_val).strip()[:10], '%Y-%m-%d').month
                except ValueError:
                    continue

        if month is None:
            continue

        for ci, (ctype, mult) in col_types.items():
            if ci >= len(row) or row[ci] is None:
                continue
            try:
                val = float(str(row[ci]).replace(',', ''))
            except (ValueError, TypeError):
                continue
            monthly[month][ctype] += val * mult

    return build_energy_data(monthly)


def process_monthly_data(data_rows, date_col, col_types):
    """Process monthly summary data."""
    energy_data = []

    for row in data_rows:
        record = {}
        month_val = 0

        # Extract month
        if date_col is not None and date_col < len(row):
            val = row[date_col]
            if val is not None:
                val_str = str(val).strip().replace(',', '')
                m = re.match(r'(\d+)', val_str)
                if m:
                    month_val = int(m.group(1))
                elif isinstance(val, (int, float)):
                    month_val = int(val)

        if month_val and 1 <= month_val <= 12:
            record['month'] = month_val

        # Extract energy values
        for ci, (ctype, mult) in col_types.items():
            if ci < len(row) and row[ci] is not None:
                try:
                    val = float(str(row[ci]).replace(',', ''))
                    record[ctype] = round(val * mult, 2)
                except (ValueError, TypeError):
                    pass

        if record:
            energy_data.append(record)

    return energy_data


def build_energy_data(monthly_dict):
    """Convert monthly defaultdict to sorted list."""
    energy_data = []
    for month in sorted(monthly_dict.keys()):
        record = {'month': int(month)}
        for ctype, val in monthly_dict[month].items():
            record[ctype] = round(val, 2)
        energy_data.append(record)
    return energy_data


def extract_building_info(meta_rows, source):
    """Extract building info from metadata rows."""
    info = {"name": "", "area": 0, "type": "", "location": "", "year": 0, "source": source}

    for row in meta_rows:
        text = ' '.join(str(c) for c in row if c is not None).strip()
        if not text:
            continue

        kv_match = re.match(r'(.+?)[:：]\s*(.+)', text)
        if kv_match:
            key = kv_match.group(1).strip()
            value = kv_match.group(2).strip()
            if any(kw in key for kw in ['建筑名称', '项目名称', '名称', '楼宇']):
                info['name'] = value
            elif any(kw in key for kw in ['建筑面积', '面积']):
                m = re.search(r'(\d+(?:\.\d+)?)', value)
                if m:
                    info['area'] = float(m.group(1))
            elif any(kw in key for kw in ['建筑类型', '类型', '用途', '功能']):
                for t in ['办公', '商业', '商场', '酒店', '宾馆', '医院', '学校', '住宅', '居住', '综合']:
                    if t in value:
                        info['type'] = t
                        break
                if not info['type']:
                    info['type'] = value
            elif any(kw in key for kw in ['地点', '位置', '城市', '省份', '地区', '地址']):
                info['location'] = value
            elif any(kw in key for kw in ['年份', '年度', '数据年份']):
                m = re.search(r'(\d{4})', value)
                if m:
                    info['year'] = int(m.group(1))

    # If name not found, try file-based
    if not info['name'] and os.path.isfile(source):
        basename = os.path.splitext(os.path.basename(source))[0]
        info['name'] = basename

    return info


# ============================================================
# Text/CSV parsing (unchanged core logic, enhanced)
# ============================================================

def parse_csv_file(filepath):
    rows = []
    with open(filepath, 'r', encoding='utf-8-sig') as f:
        reader = csv.reader(f)
        for row in reader:
            rows.append(list(row))
    return rows_to_data(rows, filepath)


def parse_table_text(text):
    lines = text.strip().split('\n')
    meta_lines, table_lines = [], []
    in_table = False
    for line in lines:
        stripped = line.strip()
        if stripped.startswith('|'):
            in_table = True
            table_lines.append(line)
        elif in_table:
            table_lines.append(line)
        else:
            meta_lines.append(line)

    if any('|' in line for line in table_lines):
        return parse_markdown_table(table_lines, meta_lines)
    if not table_lines:
        return {"building_info": {}, "energy_data": [], "error": "未找到表格数据"}
    return parse_plain_table(table_lines, meta_lines)


def parse_markdown_table(lines, meta_lines=None):
    rows = []
    for line in lines:
        line = line.strip()
        if not line or line.startswith('|--') or line.startswith('|:'):
            continue
        if line.startswith('|'):
            cells = [cell.strip() for cell in line.split('|')[1:-1]]
            rows.append(cells)
    ml = [[r] if isinstance(r, str) else list(r) for r in (meta_lines or [])]
    return rows_to_data(rows, 'table text', meta_lines=ml)


def parse_plain_table(lines, meta_lines=None):
    rows = []
    for line in lines:
        line = line.strip()
        if not line:
            continue
        cells = re.split(r'\s{2,}|\t', line)
        cells = [c.strip() for c in cells if c.strip()]
        if cells:
            rows.append(cells)
    ml = [[r] if isinstance(r, str) else list(r) for r in (meta_lines or [])]
    return rows_to_data(rows, 'table text', meta_lines=ml)


def rows_to_data(rows, source, meta_lines=None):
    if meta_lines is None:
        meta_lines = []
    if not rows or len(rows) < 2:
        return {"error": "数据不足，至少需要表头行和数据行", "building_info": {}, "energy_data": []}

    all_rows = [r for r in rows if any(c != '' and c is not None for c in r)]
    header_row_idx = find_header_row_idx_from_simple(all_rows)
    if header_row_idx is None:
        header_row_idx = 0

    headers = [normalize_header(str(h)) for h in all_rows[header_row_idx]]
    data_rows = all_rows[header_row_idx + 1:]

    all_meta = meta_lines + all_rows[:header_row_idx]
    building_info = extract_building_info(all_meta, source)

    energy_data = []
    for row in data_rows:
        record = {}
        for i, header in enumerate(headers):
            if i < len(row) and row[i] is not None:
                val_str = str(row[i]).strip().replace(',', '')
                if header == 'month':
                    m = re.match(r'(\d+)', val_str)
                    record[header] = int(m.group(1)) if m else 0
                elif header in ('electricity_kwh', 'gas_m3', 'heat_gj', 'water_ton'):
                    try:
                        record[header] = float(val_str)
                    except ValueError:
                        pass
                else:
                    try:
                        record[header] = float(val_str)
                    except ValueError:
                        pass
        if record:
            energy_data.append(record)

    return {"building_info": building_info, "energy_data": energy_data}


def find_header_row_idx_from_simple(rows):
    for i, row in enumerate(rows):
        row_text = ' '.join(str(c) for c in row if c is not None).lower()
        if any(kw in row_text for kw in ['电', '气', '月份', 'month', 'kwh', 'm³', 'gj']):
            if is_header_row(list(row)):
                return i
    return 0


def normalize_header(header):
    mapping = {
        '月份': 'month', '月': 'month', 'month': 'month',
        '用电量': 'electricity_kwh', '电力': 'electricity_kwh', '电耗': 'electricity_kwh',
        '用电量(kwh)': 'electricity_kwh', '用电量（kwh）': 'electricity_kwh',
        'electricity': 'electricity_kwh', '电量': 'electricity_kwh',
        '用气量': 'gas_m3', '天然气': 'gas_m3', '气耗': 'gas_m3',
        '用气量(m³)': 'gas_m3', '用气量（m³）': 'gas_m3',
        'gas': 'gas_m3', '燃气': 'gas_m3',
        '用热量': 'heat_gj', '热力': 'heat_gj', '热耗': 'heat_gj',
        '用热量(gj)': 'heat_gj', '用热量（gj）': 'heat_gj',
        'heat': 'heat_gj', '供热': 'heat_gj',
        '用水量': 'water_ton', '水': 'water_ton',
    }
    h_clean = str(header).strip().lower().replace(' ', '').replace('_', '')
    for key, value in mapping.items():
        if key.replace(' ', '').replace('_', '') == h_clean:
            return value
    return str(header).strip()


# ============================================================
# Main
# ============================================================

def preview_excel(filepath, max_rows=5, max_cols=30):
    """预览 Excel 文件结构，帮助用户指定列映射。"""
    try:
        import openpyxl
    except ImportError:
        return {"error": "需要安装 openpyxl: pip install openpyxl"}
    wb = openpyxl.load_workbook(filepath, data_only=True)
    result = {'sheets': {}, 'active_sheet': wb.active.title}
    for sname in wb.sheetnames:
        ws = wb[sname]
        rows = []
        for ri, row in enumerate(ws.iter_rows(min_row=1, max_row=min(max_rows, ws.max_row),
                                               max_col=min(max_cols, ws.max_column), values_only=True), 1):
            cells = {ci+1: str(c) if c is not None else '' for ci, c in enumerate(row) if c is not None}
            if cells:
                rows.append({'row': ri, 'cells': cells})
        result['sheets'][sname] = {
            'total_rows': ws.max_row, 'total_cols': ws.max_column,
            'preview_rows': rows,
        }
    wb.close()
    return result


def parse_with_columns(filepath, col_map, daily=False):
    """
    使用用户指定的列映射解析文件。
    col_map: {'electricity_kwh': 'E', 'gas_m3': 'G,Q', 'heat_gj': None}
    列标识支持字母(Excel)或数字(CSV)。
    """
    import openpyxl
    wb = openpyxl.load_workbook(filepath, data_only=True)
    result = {'building_info': {'source': filepath}, 'energy_data': []}

    # Convert column letters to 1-based indices
    def col_to_idx(col_ref):
        if isinstance(col_ref, int):
            return col_ref
        col_ref = str(col_ref).upper().strip()
        if col_ref.isdigit():
            return int(col_ref)
        result_idx = 0
        for char in col_ref:
            result_idx = result_idx * 26 + (ord(char) - ord('A') + 1)
        return result_idx

    # Parse column map
    parsed_map = {}
    for ctype, col_spec in col_map.items():
        if not col_spec:
            continue
        indices = []
        for ref in str(col_spec).split(','):
            ref = ref.strip()
            if ref:
                indices.append(col_to_idx(ref))
        if indices:
            parsed_map[ctype] = indices

    if not parsed_map:
        return {"error": "未指定有效的列映射", "building_info": {}, "energy_data": []}

    # Process first sheet with data
    for sname in wb.sheetnames:
        ws = wb[sname]
        if ws.max_row < 3:
            continue

        if daily:
            # Check if columns are cumulative or pre-calculated consumption
            is_consumption = True  # Default: assume consumption values
            for ci_list in parsed_map.values():
                for ci in ci_list:
                    header = str(ws.cell(row=2, column=ci).value or '')
                    if '用量' in header:
                        is_consumption = True
                    elif any(kw in header for kw in ['电', '气', '水']) and '用量' not in header:
                        is_consumption = False

            monthly = {}
            # Detect date issue: if first data row has year < 2000, dates are Excel serial
            first_dt = ws.cell(row=3, column=1).value
            bad_dates = isinstance(first_dt, datetime) and first_dt.year < 2000
            if bad_dates:
                # Use row index to infer month (assume ~30 rows per month)
                rows_per_month_approx = (ws.max_row - 2) / 12

            if is_consumption:
                # Pre-calculated daily consumption: just sum
                row_count = 0
                for ri in range(3, ws.max_row + 1):
                    dt_val = ws.cell(row=ri, column=1).value
                    if bad_dates:
                        row_count += 1
                        month = int((row_count - 1) / rows_per_month_approx) + 1
                        if month > 12:
                            month = 12
                    else:
                        month = None
                        if isinstance(dt_val, datetime):
                            month = dt_val.month
                        if month is None:
                            continue
                    if month not in monthly:
                        monthly[month] = {ct: 0 for ct in parsed_map}
                    for ctype, col_indices in parsed_map.items():
                        for ci in col_indices:
                            v = ws.cell(row=ri, column=ci).value
                            if v is not None:
                                try:
                                    monthly[month][ctype] += float(str(v).replace(',', ''))
                                except (ValueError, TypeError):
                                    pass
            else:
                # Cumulative meter readings: subtract previous
                initial = {}
                for ci in range(1, ws.max_column + 1):
                    v = ws.cell(row=1, column=ci).value
                    if v is not None:
                        try:
                            initial[ci] = float(str(v).replace(',', ''))
                        except (ValueError, TypeError):
                            pass
                for ri in range(3, ws.max_row + 1):
                    dt_val = ws.cell(row=ri, column=1).value
                    month = None
                    if isinstance(dt_val, datetime):
                        month = dt_val.month
                    if month is None:
                        continue
                    if month not in monthly:
                        monthly[month] = {ct: 0 for ct in parsed_map}
                    for ctype, col_indices in parsed_map.items():
                        total = 0
                        for ci in col_indices:
                            v = ws.cell(row=ri, column=ci).value
                            if v is not None:
                                try:
                                    current = float(str(v).replace(',', ''))
                                    prev = initial.get(ci, current)
                                    consumption = max(0, current - prev)
                                    initial[ci] = current
                                    total += consumption
                                except (ValueError, TypeError):
                                    pass
                        monthly[month][ctype] += round(total, 2)

            result['energy_data'] = [
                {'month': m, **vals} for m, vals in sorted(monthly.items())
            ]
        else:
            # Monthly: header row, then data rows
            monthly = []
            header_row = 1
            for ri in range(1, min(5, ws.max_row + 1)):
                row_vals = [ws.cell(row=ri, column=ci).value for ci in range(1, min(10, ws.max_column + 1))]
                text_count = sum(1 for v in row_vals if isinstance(v, str))
                if text_count >= 2:
                    header_row = ri
                    break

            for ri in range(header_row + 1, ws.max_row + 1):
                record = {}
                month_val = ws.cell(row=ri, column=1).value
                if month_val:
                    m = re.match(r'(\d+)', str(month_val))
                    if m:
                        record['month'] = int(m.group(1))

                for ctype, col_indices in parsed_map.items():
                    total = 0
                    for ci in col_indices:
                        v = ws.cell(row=ri, column=ci).value
                        if v is not None:
                            try:
                                total += float(str(v).replace(',', ''))
                            except (ValueError, TypeError):
                                pass
                    if total > 0:
                        record[ctype] = round(total, 2)

                if record.get('month'):
                    monthly.append(record)

            result['energy_data'] = sorted(monthly, key=lambda x: x.get('month', 0))

        if result['energy_data']:
            break

    wb.close()
    return result


def main():
    # Check for special flags
    args = sys.argv[1:]
    if '--preview' in args:
        idx = args.index('--preview')
        filepath = args[idx + 1] if idx + 1 < len(args) else args[0] if args else ''
        if os.path.isfile(filepath):
            result = preview_excel(filepath)
        else:
            result = {"error": f"文件不存在: {filepath}"}
        print(json.dumps(result, ensure_ascii=False, indent=2))
        return

    if '--columns' in args:
        idx = args.index('--columns')
        col_spec = args[idx + 1] if idx + 1 < len(args) else ''
        # Find filepath: skip flags and their values; filepath = remaining arg that exists on disk
        flag_with_value = {'--columns', '--preview'}
        skip_next = False
        filepath = ''
        for a in args:
            if skip_next:
                skip_next = False
                continue
            if a in flag_with_value:
                skip_next = True
                continue
            if a.startswith('--'):
                continue
            # Check if it looks like a file path (exists or has extension/drive letter)
            if os.path.exists(a) or re.match(r'^[A-Za-z]:[\\/]', a) or a.endswith(('.xlsx', '.xls', '.csv', '.txt')):
                filepath = a
        if not filepath:
            print(json.dumps({"error": "未指定文件路径。用法: parse_data.py <文件路径> --columns <列映射> [--daily]"}, ensure_ascii=False))
            sys.exit(1)

        # Parse col_spec like "E:elec,G:gas"
        col_map = {}
        for part in col_spec.split(','):
            if ':' in part:
                col_ref, ctype = part.split(':', 1)
                col_ref = col_ref.strip()
                ctype = ctype.strip()
                if ctype in col_map:
                    col_map[ctype] += ',' + col_ref
                else:
                    col_map[ctype] = col_ref
            elif '=' in part:
                col_ref, ctype = part.split('=', 1)
                col_ref = col_ref.strip()
                ctype = ctype.strip()
                if ctype in col_map:
                    col_map[ctype] += ',' + col_ref
                else:
                    col_map[ctype] = col_ref

        daily = '--daily' in args

        if os.path.isfile(filepath):
            result = parse_with_columns(filepath, col_map, daily=daily)
        else:
            result = {"error": f"文件不存在: {filepath}", "building_info": {}, "energy_data": []}
        print(json.dumps(result, ensure_ascii=False, indent=2))
        return

    # Normal mode
    if len(args) < 1:
        input_text = sys.stdin.read().strip()
        if not input_text:
            print(json.dumps({"error": "用法: parse_data.py <文件路径> [--preview <path>] [--columns <spec>] [--daily]"}, ensure_ascii=False))
            sys.exit(1)
    else:
        input_text = args[0]

    input_type, content = detect_input_type(input_text)

    try:
        if input_type == 'excel':
            result = parse_excel(content)
        elif input_type == 'csv':
            result = parse_csv_file(content)
        else:
            result = parse_table_text(content)

        print(json.dumps(result, ensure_ascii=False, indent=2))
    except Exception as e:
        import traceback
        print(json.dumps({
            "error": str(e),
            "traceback": traceback.format_exc(),
            "building_info": {}, "energy_data": [],
            "hint": "自动解析失败，请用 --preview 预览文件结构，再用 --columns 手动指定列映射"
        }, ensure_ascii=False, indent=2))
        sys.exit(1)


if __name__ == '__main__':
    main()
